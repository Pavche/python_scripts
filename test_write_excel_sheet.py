# Opens an Excel table and shows the contents of sheet with name "Sheet1"
# Using Python module openpyxl. It works with XLSX format.

import openpyxl

file_name='/home/georgiev/Zasuvky_NETIO4.xlsx'

wb = openpyxl.load_workbook(file_name)
sheet = wb.get_sheet_by_name('Sheet1')

# Reading 17 rows x 7 columns
# Python makes loop starting counter from zero with open right range
# When writing a loop with range(10) will repeat 10 times.
# When writing a loop with range(1,10) will repeat 9 times.

for i in range(1,18):
    for j in range(1,8):
        print(sheet.cell(row=i,column=j).value)

print('Highest row in Excel sheet:',sheet.max_row)
print('Highest column in Excel sheet:',sheet.max_column)

exit()