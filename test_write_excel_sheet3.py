# Scenario
# Create an Excel workbook with one sheet which contains 7 fields (columns)
# 1st - Location
# 2nd - Host name
# 3rd - IP address
# 4th - MAC address
# 5th - VLAN
# 6th - Switch name
# 7th - Port name
# On row No.5 all cells have bold and italic font, aligment_horiz=Center, aligment_vert=Center, font color=BLUE

# Cell formatting in Excel sheet are shown here:
# https://openpyxl.readthedocs.org/en/default/styles.html

import openpyxl
from openpyxl.styles import Font, Style, Alignment

XLSX_FILE='/home/georgiev/test_host_list.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Host list"

header=['Location','Host name','IP address','MAC address','VLAN','Switch name','Port name']
col=1
for caption in header:
    c = ws.cell(row=5,column=col)
    c.value = caption
    c.font  = Font(size=14, color=openpyxl.styles.colors.BLUE, italic=True, bold=True)
    c.alignment = Alignment(horizontal='center',vertical='center')
    col+=1

# Open a text file in read-only mode.
# The text file has 7 columns, each separated by tab.
# Read the it line by line.
# From 6th row on the sheet put the contents of the text file
# Put one line of text on each row.
# Every word is put in separate cell on the sheet.
# Word are sparated by tab.


TXT_FILE = '/home/georgiev/host_list.in'
f = open(TXT_FILE,'r')

my_row=6
my_col=1
pad=""

#for line in f:
#    print line.strip()

for line in f:    
    # ws.cell(row=my_row, column=my_col).value = line.strip()
    pad = line.strip()
    
    # How to create from one line a list of words? Use lists with string functions.
    # string.split(s[, sep[, maxsplit]])
    list_pad=pad.split()
    
    # Writing the contents into one row in the Excel sheet
    for element in list_pad:
        ws.cell(row=my_row, column=my_col).value = element
        my_col+=1
    
    my_row+=1 # move to the next row in the sheet
    my_col=1 # return to the fisrt (initial) column in the sheet
    

f.close()



# Save the Excel workbook
wb.save(XLSX_FILE)
print 'Excel workbook '+XLSX_FILE+' was created.'
