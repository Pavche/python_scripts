import openpyxl

XLSX_FILE='/home/georgiev/test_dhcp_cfg.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Network IDs"
message = 'Fist steps in creating Excel workbooks with Python'
ws['A1'] = message
#For cell styles and formating check the documentation:
#https://openpyxl.readthedocs.org/en/default/styles.html
A1 = ws['A1']
A1.font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE, italic=True)


# Printing from 1 to 12 in column A
for i in range(1,13):
    ws.cell(row=i+3, column=1).value = i

wb.save(XLSX_FILE)
print 'Excel workbook '+XLSX_FILE+' was created.'
#except:
#    print 'Problem when creating a new Excel workbook.'